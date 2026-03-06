"""
app.py — Tra cứu MST TNCN hàng loạt qua masothue.com
Dùng: python app.py
      python app.py --input ds.csv --output ketqua.xlsx
      python app.py --mst 079203002600
      python app.py --format csv
      python app.py --workers 5  # Chạy song song với 5 browsers
      python app.py --no-parallel  # Tắt chế độ song song
"""

import argparse
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from scraper import lookup_batch, TaxRecord


def load_file(path: str) -> list[dict]:
    """Load CSV hoặc Excel, tự động tìm cột MST"""
    ext = path.lower().rsplit('.', 1)[-1]
    
    # Đọc file
    try:
        if ext == 'csv':
            # Thử nhiều encoding cho CSV
            for encoding in ['utf-8', 'utf-8-sig', 'latin1', 'cp1252', 'iso-8859-1']:
                try:
                    df = pd.read_csv(path, dtype=str, encoding=encoding).fillna("")
                    break
                except UnicodeDecodeError:
                    continue
            else:
                print(f"❌ Không thể đọc file CSV với các encoding thông dụng")
                sys.exit(1)
        elif ext in ('xlsx', 'xls'):
            # Excel không cần encoding, pandas tự xử lý
            df = pd.read_excel(path, dtype=str, engine='openpyxl' if ext == 'xlsx' else None).fillna("")
        else:
            print(f"❌ Định dạng file không hỗ trợ: {ext}")
            print("   Chỉ hỗ trợ: .csv, .xlsx, .xls")
            sys.exit(1)
    except Exception as e:
        print(f"❌ Lỗi đọc file: {str(e)}")
        sys.exit(1)
    
    # Tìm và map các cột
    col_map = {}
    for col in df.columns:
        c = col.strip().lower()
        if c in ("mst", "ma_so_thue", "masothue", "ma so thue", "tax_id", "cccd", "cmnd"):
            col_map[col] = "mst"
        elif c in ("ho_ten", "hoten", "ho ten", "fullname", "ten", "name"):
            col_map[col] = "ho_ten"
        elif c in ("dia_chi", "diachi", "dia chi", "address", "dc"):
            col_map[col] = "dia_chi"
    
    df.rename(columns=col_map, inplace=True)
    
    # Kiểm tra cột MST
    if "mst" not in df.columns:
        print("❌ Không tìm thấy cột MST trong file")
        print(f"   Các cột hiện có: {', '.join(df.columns)}")
        print("   Vui lòng đặt tên cột là: MST, ma_so_thue, CCCD, hoặc CMND")
        sys.exit(1)
    
    # Lọc các dòng có MST
    records = []
    for _, row in df.iterrows():
        mst = str(row.get("mst", "")).strip()
        if mst and mst.lower() not in ('nan', 'none', ''):
            records.append({
                "mst": mst,
                "ho_ten": str(row.get("ho_ten", "")).strip(),
                "dia_chi": str(row.get("dia_chi", "")).strip()
            })
    
    if not records:
        print("❌ Không tìm thấy MST nào trong file")
        sys.exit(1)
    
    print(f"✅ Đọc được {len(records)} MST từ {path}")
    return records


HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"),
)
COLUMNS = [
    ("STT", 6), ("MST / CCCD (nhập)", 18), ("Họ và tên (nhập)", 22),
    ("Tên NNT (kết quả)", 26), ("MST (kết quả)", 16),
    ("Địa chỉ (kết quả)", 30), ("Ghi chú / Lỗi", 24),
]


def export_excel(records: list[TaxRecord], path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Kết quả tra cứu"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"KẾT QUẢ TRA CỨU MST/CCCD — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79", name="Arial")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=2, column=ci, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[2].height = 30

    OK_FILL  = PatternFill("solid", start_color="E2EFDA")
    ALT_FILL = PatternFill("solid", start_color="D6E4F0")
    ERR_FILL = PatternFill("solid", start_color="FCE4D6")

    for i, r in enumerate(records, 1):
        row = i + 2
        fill = ERR_FILL if r.loi else (ALT_FILL if i % 2 == 0 else OK_FILL)
        for ci, val in enumerate([
            i, r.mst, r.ho_ten_input,
            r.ten_nnt, r.mst_result,
            r.dia_chi_result, r.loi,
        ], 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill = fill
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = BORDER
        ws.row_dimensions[row].height = 18

    ws.freeze_panes = "A3"

    sr = len(records) + 4
    ok = sum(1 for r in records if r.ten_nnt)
    for label, val, color in [
        ("Tổng:", len(records), "000000"),
        ("Tìm thấy:", ok, "375623"),
        ("Không tìm thấy:", len(records) - ok, "C00000"),
    ]:
        ws.cell(row=sr, column=1, value=label).font = Font(bold=True, color=color, name="Arial")
        ws.cell(row=sr, column=2, value=val)
        sr += 1

    wb.save(path)
    print(f"✅ Đã xuất Excel: {path}")


def export_csv(records: list[TaxRecord], path: str):
    pd.DataFrame([{
        "STT": i + 1, "MST/CCCD (nhập)": r.mst, "Họ tên (nhập)": r.ho_ten_input,
        "Tên NNT": r.ten_nnt, "MST (kết quả)": r.mst_result,
        "Địa chỉ": r.dia_chi_result, "Lỗi": r.loi,
    } for i, r in enumerate(records)]).to_csv(path, index=False, encoding="utf-8-sig")
    print(f"✅ Đã xuất CSV: {path}")


def main():
    parser = argparse.ArgumentParser(
        description="Tra cứu MST/CCCD hàng loạt",
        epilog="Ví dụ:\n"
               "  python app.py --input data.csv\n"
               "  python app.py --input data.xlsx --output result.xlsx\n"
               "  python app.py --mst 0123456789\n"
               "  python app.py --workers 5\n",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--input",  default="data.csv", help="File CSV/Excel đầu vào (mặc định: data.csv)")
    parser.add_argument("--output", default="output.xlsx", help="File kết quả (mặc định: output.xlsx)")
    parser.add_argument("--format", choices=["xlsx", "csv"], help="Định dạng output (tự động từ tên file nếu không chỉ định)")
    parser.add_argument("--mst",    help="Tra cứu nhanh 1 MST/CCCD")
    parser.add_argument("--workers", type=int, help="Số lượng browser workers (1-10, mặc định từ config)")
    parser.add_argument("--no-parallel", action="store_true", help="Tắt chế độ song song")
    args = parser.parse_args()

    records_input = [{"mst": args.mst}] if args.mst else load_file(args.input)
    if not records_input:
        print("❌ Không có dữ liệu")
        sys.exit(1)

    # Xác định parallel mode
    parallel = None
    if args.no_parallel:
        parallel = False
    elif args.workers:
        parallel = args.workers
    
    print(f"\n🔍 Bắt đầu tra cứu {len(records_input)} MST...\n")
    
    # Import scraper factory
    from scraper_factory import get_scraper
    
    # Sử dụng context manager để tự động đóng browser
    with get_scraper(source="tax-3rd", headless=True, parallel=parallel) as scraper:
        results = scraper.lookup_batch(records_input)

    # Xác định format output
    output_format = args.format
    if not output_format:
        # Tự động từ extension
        output_format = "csv" if args.output.endswith(".csv") else "xlsx"
    
    if args.format == "csv" or args.output.endswith(".csv"):
        export_csv(results, args.output)
    else:
        export_excel(results, args.output)

    ok = sum(1 for r in results if r.ten_nnt)
    print(f"\n📊 Hoàn tất: {ok}/{len(results)} tìm thấy thông tin")


if __name__ == "__main__":
    main()