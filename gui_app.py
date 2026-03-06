"""
GUI App for Tax Lookup - Redesigned UI
Tất cả nút chức năng ở header, mặc định Cục Thuế, có nút Dừng
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from pathlib import Path
import threading
from scraper_factory import get_scraper, get_source_info

class TaxLookupGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Tra cứu Mã số thuế - Cục Thuế & MaSoThue.com")
        self.root.geometry("1400x800")
        
        # Data storage
        self.input_data = []
        self.result_data = []
        self.is_searching = False
        self.stop_requested = False
        self.current_scraper = None
        
        # Settings
        self.selected_source = "tax-gov"  # Mặc định: Cục Thuế
        self.headless_var = tk.BooleanVar(value=True)
        self.parallel_var = tk.BooleanVar(value=False)
        self.num_workers_var = tk.IntVar(value=1)
        
        self.setup_ui()
        self.on_source_change()
        
    def setup_ui(self):
        # Main container
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # Header
        self.setup_header(main_frame)
        
        # Table
        self.setup_table(main_frame)
        
    def setup_header(self, parent):
        header = ttk.Frame(parent, padding="10")
        header.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # Row 1: Nguồn + File operations
        r1 = ttk.Frame(header)
        r1.pack(fill=tk.X, pady=(0, 12))
        
        ttk.Label(r1, text="Nguồn:", font=("", 11, "bold")).pack(side=tk.LEFT, padx=(0, 10))
        
        self.source_var = tk.StringVar(value="tax-gov")
        source_info = get_source_info()
        
        ttk.Radiobutton(r1, text=f"🏛️ {source_info['tax-gov']['name']}", 
                       variable=self.source_var, value="tax-gov",
                       command=self.on_source_change).pack(side=tk.LEFT, padx=5)
        
        ttk.Radiobutton(r1, text=f"🌐 {source_info['tax-3rd']['name']}", 
                       variable=self.source_var, value="tax-3rd",
                       command=self.on_source_change).pack(side=tk.LEFT, padx=5)
        
        ttk.Separator(r1, orient='vertical').pack(side=tk.LEFT, fill=tk.Y, padx=20)
        
        ttk.Button(r1, text="📂 Load CSV/Excel", command=self.load_csv, width=18).pack(side=tk.LEFT, padx=5)
        ttk.Button(r1, text="🗑️ Xóa dữ liệu", command=self.clear_data, width=15).pack(side=tk.LEFT, padx=5)
        
        self.export_btn = ttk.Button(r1, text="💾 Tải Excel", command=self.export_excel, 
                                     state=tk.DISABLED, width=15)
        self.export_btn.pack(side=tk.LEFT, padx=5)
        
        # Row 2: Nhập thủ công
        r2 = ttk.Frame(header)
        r2.pack(fill=tk.X, pady=(0, 12))
        
        ttk.Label(r2, text="Nhập thủ công:", font=("", 10, "bold")).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Label(r2, text="MST:").pack(side=tk.LEFT, padx=(0, 5))
        self.mst_entry = ttk.Entry(r2, width=15)
        self.mst_entry.pack(side=tk.LEFT, padx=(0, 15))
        
        self.cccd_label = ttk.Label(r2, text="CCCD:")
        self.cccd_entry = ttk.Entry(r2, width=15)
        
        self.name_label = ttk.Label(r2, text="Họ tên:")
        self.name_label.pack(side=tk.LEFT, padx=(0, 5))
        self.name_entry = ttk.Entry(r2, width=20)
        self.name_entry.pack(side=tk.LEFT, padx=(0, 15))
        
        self.address_label = ttk.Label(r2, text="Địa chỉ:")
        self.address_entry = ttk.Entry(r2, width=25)
        
        ttk.Button(r2, text="➕ Thêm", command=self.add_manual_entry, width=10).pack(side=tk.LEFT, padx=5)
        
        # Row 3: Settings + Actions
        r3 = ttk.Frame(header)
        r3.pack(fill=tk.X, pady=(0, 12))
        
        ttk.Checkbutton(r3, text="🔇 Chạy ẩn", variable=self.headless_var).pack(side=tk.LEFT, padx=(0, 15))
        
        ttk.Checkbutton(r3, text="⚡ Song song", variable=self.parallel_var,
                       command=self.on_parallel_toggle).pack(side=tk.LEFT, padx=(0, 10))
        
        self.workers_label = ttk.Label(r3, text="Workers: 1")
        self.workers_label.pack(side=tk.LEFT, padx=(0, 5))
        
        self.workers_slider = ttk.Scale(r3, from_=1, to=10, orient=tk.HORIZONTAL,
                                       variable=self.num_workers_var, command=self.on_workers_change,
                                       length=100, state=tk.DISABLED)
        self.workers_slider.pack(side=tk.LEFT, padx=(0, 20))
        
        ttk.Separator(r3, orient='vertical').pack(side=tk.LEFT, fill=tk.Y, padx=15)
        
        self.search_btn = ttk.Button(r3, text="🔍 Bắt đầu tra cứu", command=self.start_search, width=18)
        self.search_btn.pack(side=tk.LEFT, padx=5)
        
        self.stop_btn = ttk.Button(r3, text="⏹️ Dừng", command=self.stop_search, 
                                   state=tk.DISABLED, width=12)
        self.stop_btn.pack(side=tk.LEFT, padx=5)
        
        # Row 4: Status
        r4 = ttk.Frame(header)
        r4.pack(fill=tk.X)
        
        self.status_label = ttk.Label(r4, text="Sẵn sàng - Chọn file hoặc nhập MST để bắt đầu",
                                     foreground="green", font=("", 10))
        self.status_label.pack(side=tk.LEFT, padx=(0, 15))
        
        self.progress = ttk.Progressbar(r4, mode='determinate', length=350)
        self.progress.pack(side=tk.LEFT, padx=5)
        
        self.progress_label = ttk.Label(r4, text="0/0", font=("", 9))
        self.progress_label.pack(side=tk.LEFT, padx=5)
        
    def setup_table(self, parent):
        frame = ttk.LabelFrame(parent, text="Dữ liệu tra cứu", padding="10")
        frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)
        
        tree_frame = ttk.Frame(frame)
        tree_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        columns = ("STT", "MST Input", "Tên Input", "Tên KQ", "MST KQ", "Địa chỉ", "CQ Thuế", "Trạng thái", "Link")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings",
                                yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        
        widths = [50, 120, 150, 200, 120, 250, 150, 120, 80]
        for col, width in zip(columns, widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, minwidth=50)
        
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        vsb.grid(row=0, column=1, sticky=(tk.N, tk.S))
        hsb.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        self.tree.bind("<Double-1>", self.on_double_click)
        
    def on_source_change(self):
        source = self.source_var.get()
        self.selected_source = source
        
        if source == "tax-gov":
            self.cccd_label.pack_forget()
            self.cccd_entry.pack_forget()
            self.address_label.pack(side=tk.LEFT, padx=(0, 5))
            self.address_entry.pack(side=tk.LEFT, padx=(0, 15))
            # Tax-gov: Cho phép parallel với ít workers hơn (1-5)
            self.parallel_var.set(False)
            self.num_workers_var.set(1)
            self.workers_slider.config(state=tk.DISABLED, to=5)  # Max 5 workers cho tax-gov
        else:
            self.address_label.pack_forget()
            self.address_entry.pack_forget()
            self.cccd_label.pack(side=tk.LEFT, padx=(0, 5))
            self.cccd_entry.pack(side=tk.LEFT, padx=(0, 15))
            # Tax-3rd: Parallel với nhiều workers hơn (1-10)
            self.parallel_var.set(True)
            self.num_workers_var.set(5)
            self.workers_slider.config(state=tk.NORMAL, to=10)
    
    def on_parallel_toggle(self):
        state = tk.NORMAL if self.parallel_var.get() else tk.DISABLED
        self.workers_slider.config(state=state)
        
    def on_workers_change(self, value):
        num = int(float(value))
        self.workers_label.config(text=f"Workers: {num}")
        self.num_workers_var.set(num)
    
    def load_csv(self):
        filename = filedialog.askopenfilename(
            title="Chọn file CSV/Excel",
            filetypes=[("CSV/Excel files", "*.csv *.xlsx *.xls"), ("All files", "*.*")]
        )
        if not filename:
            return
            
        try:
            ext = filename.lower().rsplit('.', 1)[-1]
            
            if ext == 'csv':
                df = None
                for encoding in ['utf-8', 'utf-8-sig', 'latin1', 'cp1252', 'iso-8859-1']:
                    try:
                        df = pd.read_csv(filename, dtype=str, encoding=encoding).fillna("")
                        break
                    except UnicodeDecodeError:
                        continue
                if df is None:
                    messagebox.showerror("Lỗi", "Không thể đọc file CSV")
                    return
            elif ext in ('xlsx', 'xls'):
                df = pd.read_excel(filename, dtype=str, engine='openpyxl' if ext == 'xlsx' else None).fillna("")
            else:
                messagebox.showerror("Lỗi", f"Định dạng không hỗ trợ: {ext}")
                return
            
            self.input_data = []
            for _, row in df.iterrows():
                entry = {}
                for col in df.columns:
                    col_lower = str(col).lower().strip()
                    if 'mst' in col_lower or 'tax' in col_lower:
                        entry['mst'] = str(row[col]).strip() if pd.notna(row[col]) else ""
                    elif 'cccd' in col_lower or 'cmnd' in col_lower:
                        entry['cccd'] = str(row[col]).strip() if pd.notna(row[col]) else ""
                    elif 'ten' in col_lower or 'name' in col_lower or 'ho' in col_lower:
                        entry['ho_ten'] = str(row[col]).strip() if pd.notna(row[col]) else ""
                    elif 'dia' in col_lower or 'address' in col_lower:
                        entry['address'] = str(row[col]).strip() if pd.notna(row[col]) else ""
                
                if entry.get('mst') or entry.get('cccd'):
                    if not entry.get('mst'):
                        entry['mst'] = entry.get('cccd', '')
                    if entry['mst'] and entry['mst'].lower() not in ('nan', 'none', ''):
                        self.input_data.append(entry)
            
            if not self.input_data:
                messagebox.showwarning("Cảnh báo", "Không tìm thấy MST nào")
                return
            
            self.refresh_table()
            self.status_label.config(text=f"Đã load {len(self.input_data)} dòng", foreground="green")
            messagebox.showinfo("Thành công", f"Đã load {len(self.input_data)} dòng")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể đọc file:\n{str(e)}")
    
    def add_manual_entry(self):
        mst = self.mst_entry.get().strip()
        cccd = self.cccd_entry.get().strip() if hasattr(self, 'cccd_entry') else ""
        name = self.name_entry.get().strip()
        address = self.address_entry.get().strip() if hasattr(self, 'address_entry') else ""
        
        if self.selected_source == "tax-gov":
            if not mst:
                messagebox.showwarning("Cảnh báo", "Vui lòng nhập MST")
                return
        else:
            if not mst and not cccd:
                messagebox.showwarning("Cảnh báo", "Vui lòng nhập MST hoặc CCCD")
                return
        
        entry = {'mst': mst or cccd, 'cccd': cccd, 'ho_ten': name, 'address': address}
        self.input_data.append(entry)
        self.refresh_table()
        
        self.mst_entry.delete(0, tk.END)
        if hasattr(self, 'cccd_entry'):
            self.cccd_entry.delete(0, tk.END)
        self.name_entry.delete(0, tk.END)
        if hasattr(self, 'address_entry'):
            self.address_entry.delete(0, tk.END)
        
        self.status_label.config(text=f"Đã thêm. Tổng: {len(self.input_data)}", foreground="green")
    
    def clear_data(self):
        if not self.input_data and not self.result_data:
            return
        if messagebox.askyesno("Xác nhận", "Xóa toàn bộ dữ liệu?"):
            self.input_data = []
            self.result_data = []
            self.refresh_table()
            self.export_btn.config(state=tk.DISABLED)
            self.status_label.config(text="Đã xóa dữ liệu", foreground="gray")
    
    def refresh_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        data = self.result_data if self.result_data else self.input_data
        
        for idx, item in enumerate(data, 1):
            if hasattr(item, 'mst'):
                values = (idx, item.mst, item.ho_ten_input, item.ten_nnt, item.mst_result,
                         item.dia_chi_result[:50] + "..." if len(item.dia_chi_result) > 50 else item.dia_chi_result,
                         item.co_quan_thue, item.trang_thai or item.loi, "🔗" if item.url else "")
                self.tree.insert("", tk.END, values=values, tags=(item.url,) if item.url else ())
            elif isinstance(item, dict):
                values = (idx, item.get('mst', ''), item.get('ho_ten', ''), "", "", "", "", "", "")
                self.tree.insert("", tk.END, values=values)
    
    def on_double_click(self, event):
        import webbrowser
        selection = self.tree.selection()
        if not selection:
            return
        tags = self.tree.item(selection[0], 'tags')
        if tags and tags[0]:
            webbrowser.open(tags[0])
    
    def start_search(self):
        if not self.input_data:
            messagebox.showwarning("Cảnh báo", "Chưa có dữ liệu")
            return
        if self.is_searching:
            messagebox.showinfo("Thông báo", "Đang tra cứu...")
            return
        
        self.is_searching = True
        self.stop_requested = False
        self.search_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.progress['maximum'] = len(self.input_data)
        self.progress['value'] = 0
        self.progress_label.config(text=f"0/{len(self.input_data)}")
        self.status_label.config(text="Đang tra cứu...", foreground="orange")
        
        threading.Thread(target=self.perform_search, daemon=True).start()
    
    def stop_search(self):
        self.stop_requested = True
        self.stop_btn.config(state=tk.DISABLED)
        self.status_label.config(text="Đang dừng...", foreground="orange")
    
    def perform_search(self):
        try:
            parallel = self.num_workers_var.get() if self.parallel_var.get() else False
            headless = self.headless_var.get()
            
            # Khởi tạo result_data
            total = len(self.input_data)
            self.result_data = [None] * total
            
            # Callback để update progress từ scraper
            def progress_callback(current, total, result=None):
                self.root.after(0, lambda: self.update_progress_with_result(current, total, result))
            
            self.current_scraper = get_scraper(
                self.selected_source, 
                headless=headless, 
                parallel=parallel,
                progress_callback=progress_callback
            )
            self.current_scraper.start()
            
            # Gọi lookup_batch với TẤT CẢ items (cả single và parallel mode)
            results = self.current_scraper.lookup_batch(self.input_data)
            
            # Lưu kết quả cuối cùng (đảm bảo không có None)
            self.result_data = results
            
            # Update progress cuối cùng
            self.root.after(0, lambda: self.update_progress(total, total))
            
            self.current_scraper.close()
            self.root.after(0, self.search_complete)
            
        except Exception as e:
            self.root.after(0, lambda msg=str(e): self.search_error(msg))
    
    def update_progress(self, current, total):
        self.progress['value'] = current
        self.progress_label.config(text=f"{current}/{total}")
        self.refresh_table()
    
    def update_progress_with_result(self, current, total, result):
        """Update progress và thêm result vào danh sách ngay lập tức"""
        if result:
            # Khởi tạo result_data nếu chưa có
            if not self.result_data or len(self.result_data) != total:
                self.result_data = [None] * total
            
            # Tìm index của result dựa vào MST
            for i, inp in enumerate(self.input_data):
                if inp.get('mst') == result.mst:
                    self.result_data[i] = result
                    break
        
        self.update_progress(current, total)
    
    def search_complete(self):
        self.is_searching = False
        self.search_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        
        success = sum(1 for r in self.result_data if not r.loi)
        status = "Hoàn tất" if not self.stop_requested else "Đã dừng"
        self.status_label.config(text=f"{status}: {success}/{len(self.result_data)} thành công", foreground="green")
        self.refresh_table()
        
        # Hỏi có muốn lưu file không
        if messagebox.askyesno(status, f"Đã tra cứu {len(self.result_data)} mục.\n\nBạn có muốn lưu kết quả ngay không?"):
            self.export_excel()
        else:
            self.export_btn.config(state=tk.NORMAL)
            messagebox.showinfo("Thông báo", "Bạn có thể lưu kết quả sau bằng nút 'Tải Excel'")
    
    def search_error(self, error_msg):
        self.is_searching = False
        self.search_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.status_label.config(text="Lỗi tra cứu", foreground="red")
        messagebox.showerror("Lỗi", f"Có lỗi:\n{error_msg}")
    
    def export_excel(self):
        if not self.result_data:
            messagebox.showwarning("Cảnh báo", "Chưa có kết quả")
            return
        
        try:
            from datetime import datetime
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment
            from openpyxl.utils import get_column_letter
            
            # Tạo tên file mặc định
            date_str = datetime.now().strftime("%d%m%Y")
            default_name = f"CheckMST_Template_processed_{date_str}_1.xlsx"
            
            # Hỏi người dùng chọn nơi lưu
            filename = filedialog.asksaveasfilename(
                title="Lưu file Excel",
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not filename:
                # Người dùng hủy
                return
            
            # Load template
            template_path = Path("base/TAX_FORM.xlsx")
            if not template_path.exists():
                messagebox.showerror("Lỗi", "Không tìm thấy template tại base/TAX_FORM.xlsx")
                return
            
            # Copy template
            wb = load_workbook(template_path)
            ws = wb.active
            
            # Unmerge tất cả merged cells để có thể ghi dữ liệu
            merged_ranges = list(ws.merged_cells.ranges)
            for merged_range in merged_ranges:
                ws.unmerge_cells(str(merged_range))
            
            # Hàm viết hoa chữ cái đầu
            def title_case(text):
                if not text:
                    return ""
                # Viết hoa chữ cái đầu mỗi từ
                return ' '.join(word.capitalize() for word in text.split())
            
            # Điền dữ liệu từ dòng 2 (dòng 1 là header)
            start_row = 2
            for idx, r in enumerate(self.result_data, start=1):
                row = start_row + idx - 1
                
                # A: STT
                ws.cell(row=row, column=1).value = idx
                
                # B: Họ và Tên - Lấy từ kết quả và viết hoa chữ cái đầu
                ho_ten = title_case(r.ten_nnt) if r.ten_nnt else ""
                ws.cell(row=row, column=2).value = ho_ten
                
                # C: CMND/CCCD - Luôn lấy từ kết quả (mst_result)
                ws.cell(row=row, column=3).value = r.mst_result
                
                # D: MST 1 - Luôn lấy từ input
                mst_value = r.mst if len(r.mst) == 10 else ""
                ws.cell(row=row, column=4).value = mst_value
                
                # E: Tên người nộp thuế (kết quả gốc - giữ nguyên)
                ws.cell(row=row, column=5).value = r.ten_nnt
                
                # F: Cơ quan thuế
                ws.cell(row=row, column=6).value = r.co_quan_thue
                
                # G: Ghi chú (lỗi hoặc trạng thái)
                note = r.loi if r.loi else r.trang_thai
                ws.cell(row=row, column=7).value = note
                
                # Set alignment cho tất cả cells
                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # Lưu file
            wb.save(filename)
            
            self.status_label.config(text=f"Đã xuất: {Path(filename).name}", foreground="green")
            self.export_btn.config(state=tk.NORMAL)
            messagebox.showinfo("Thành công", f"Đã xuất file:\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xuất Excel:\n{str(e)}")


def main():
    root = tk.Tk()
    app = TaxLookupGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
